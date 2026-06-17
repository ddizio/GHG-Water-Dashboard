#!/usr/bin/env python3
"""
build_data.py — Vantage Sustainability Dashboard ETL

Reads the source Excel workbooks in the repo root and writes a single
normalized `data.json` (the dashboard's source of truth), plus a console
validation report against known anchor numbers.

Design notes:
- Pulls from the *cleanest* summary sheet for each metric (not raw site rows).
- Normalizes inconsistent site names via SITE_MASTER aliases.
- Converts all water to megalitres (ML); emissions stay in tCO2e.
- Re-run each year after dropping in the new workbook(s). See README.

Usage:  python build/build_data.py
"""
import json, os, re, sys
from datetime import date
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FILES = {
    "cdp":   "CDP Question List (2025 Update).xlsx",
    "fy24s": "VSC_S1 S2 and S3_FY2024_Inventory Summary_v2.xlsx",
    "fy24":  "VSC_ FY24_Scope1&2_Inventory_v2_06252025.xlsx",
    "fy25":  "VSC_ FY25_Scope1&2_Inventory_v2_5-20-2026 (1).xlsx",
    "s3_23": "VSC_Scope 3_FY2023_Summary Workbook.xlsx",
    "prod":  "Production Volumes - Vantage.xlsx",
    "rz":    "Copy of RouteZero_results_template_v3_VantageFY21-22.xlsx",
}

GAL_TO_ML = 3.785411784e-6   # 1 US gallon -> ML  (gal * 3.785411784 L / 1e6 L/ML)

# ---------------------------------------------------------------------------
# Site master. status: active | closed. stress/risk from WWF (Summary Water 2024).
# region/type curated. "producing" flags the 7 manufacturing sites.
# ---------------------------------------------------------------------------
SITE_MASTER = [
    # id, name, country, region, type, status, stress, risk, producing, aliases
    ("gurnee","Gurnee, IL","USA","North America","Factory","active","",  "High",   True,  ["gurnee","gurnee il"]),
    ("chicago","Chicago, IL","USA","North America","Factory","active","", "High",   True,  ["chicago","chicago il","chicago oleo"]),
    ("leuna","Leuna","Germany","Europe","Factory","active","High","High", True,  ["leuna"]),
    ("carnegie","Carnegie, PA","USA","North America","Factory","active","","Medium",True,  ["carnegie","carnegie pa","carnegie mallet ingredients"]),
    ("englewood","Englewood, OH","USA","North America","Factory","active","","High",True,  ["englewood","englewood oh"]),
    ("les_borges","Les Borges","Spain","Europe","Factory","active","High","High",  True,  ["les borges","spain"]),
    ("tucson","Tucson","USA","North America","Factory","active","High","High",     True,  ["tucson"]),
    ("agrinsa","Agrinsa (Argentina Farm)","Argentina","South America","Farm","active","","Medium",False,["agrinsa","argentina farm agrinsa","argentina farm"]),
    ("dateland","Dateland","USA","North America","Farm","active","High","High",    False, ["dateland"]),
    ("barcelona","Barcelona","Spain","Europe","Office","active","","",             False, ["barcelona"]),
    ("granollers","Granollers","Spain","Europe","Office","closed","","",           False, ["granollers"]),
    ("mexico_city","Mexico City","Mexico","North America","Warehouse","active","","",False,["mexico city"]),
    ("linden","Linden, NJ","USA","North America","Warehouse","closed","","",       False, ["linden","linden nj"]),
    ("fairfield","Fairfield","USA","North America","Warehouse","active","High","Medium",False,["fairfield"]),
    ("warren","Warren, NJ","USA","North America","Office","active","","",          False, ["warren","warren nj"]),
    ("pittsburgh","Pittsburgh, PA","USA","North America","Office","active","","",   False, ["pittsburgh","pittsburgh pa"]),
    ("south_africa","South Africa","South Africa","Africa","Warehouse","active","","",False,["south africa"]),
    ("delta","Delta","USA","North America","Warehouse","closed","","",             False, ["delta"]),
    ("deerfield","Deerfield","USA","North America","Office","active","","",         False, ["deerfield"]),
    ("tlalnepantla","Tlalnepantla de Baz","Mexico","North America","Warehouse","closed","","",False,["tlalnepantla de baz","tlalnepantla"]),
    ("shanghai","Shanghai","China","Asia","Warehouse","active","","",             False, ["shanghai"]),
    ("bogota","Bogota","Colombia","South America","Office","active","","",         False, ["bogota","bogotá"]),
    ("santiago","Santiago","Chile","South America","Warehouse","active","","",     False, ["santiago"]),
    ("guatemala","Guatemala City","Guatemala","North America","Warehouse","closed","","",False,["guatemala city"]),
    ("sao_paulo","Sao Paulo","Brazil","South America","Warehouse","active","","",  False, ["sao paulo","são paulo"]),
    ("buenos_aires","Buenos Aires","Argentina","South America","Office","active","","",False,["buenos aires"]),
    ("mumbai","Mumbai","India","Asia","Warehouse","active","","",                 False, ["mumbai"]),
    ("lima","Lima","Peru","South America","Warehouse","closed","","",             False, ["lima"]),
    ("costa_rica","Costa Rica","Costa Rica","North America","Warehouse","closed","","",False,["costa rica","costa_rica"]),
    ("paris","Paris","France","Europe","Office","closed","","",                   False, ["paris"]),
]

def norm(s):
    if s is None: return ""
    s = str(s).strip().lower()
    s = s.replace("vantage_", "").replace("vantage ", "")
    s = re.sub(r"[\(\),\.\-_/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

ALIAS = {}
for row in SITE_MASTER:
    sid, name, *_rest, aliases = row
    ALIAS[norm(name)] = sid
    ALIAS[norm(sid)] = sid
    for a in aliases:
        ALIAS[norm(a)] = sid

UNMAPPED = set()
_NON_SITE = {"column1","production volume","total","grand total","site","totals","year","scope"}
def site_id(name):
    n = norm(name)
    if n in ALIAS: return ALIAS[n]
    # try without trailing tokens like "oleo", "mallet ingredients"
    for key in ALIAS:
        if n.startswith(key) or key.startswith(n):
            return ALIAS[key]
    # ignore obvious header/title/numeric cells rather than flag them
    if n and n not in _NON_SITE and not re.fullmatch(r"[\d ]+", n) and len(n) < 40:
        UNMAPPED.add(str(name))
    return None

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(",", "").strip()
    try: return float(s)
    except: return None

def load(key):
    return openpyxl.load_workbook(os.path.join(ROOT, FILES[key]), read_only=True, data_only=True)

def rows(ws, maxr=None, maxc=None):
    return list(ws.iter_rows(min_row=1, max_row=maxr or ws.max_row,
                             max_col=maxc or ws.max_column, values_only=True))

# Site coordinates [lat, lon] for the map (approx city/site locations)
COORDS = {
 "gurnee":[42.37,-87.92],"chicago":[41.85,-87.65],"leuna":[51.32,12.01],
 "carnegie":[40.41,-80.08],"englewood":[39.88,-84.30],"les_borges":[41.52,0.87],
 "tucson":[32.22,-110.97],"agrinsa":[-34.0,-60.5],"dateland":[32.80,-113.54],
 "barcelona":[41.39,2.17],"granollers":[41.61,2.29],"mexico_city":[19.43,-99.13],
 "linden":[40.62,-74.24],"fairfield":[40.88,-74.30],"warren":[40.63,-74.51],
 "pittsburgh":[40.44,-79.99],"south_africa":[-26.20,28.04],"deerfield":[42.17,-87.84],
 "tlalnepantla":[19.54,-99.20],"shanghai":[31.23,121.47],"bogota":[4.71,-74.07],
 "santiago":[-33.45,-70.67],"guatemala":[14.63,-90.51],"sao_paulo":[-23.55,-46.63],
 "buenos_aires":[-34.60,-58.38],"mumbai":[19.08,72.88],"lima":[-12.05,-77.04],
 "costa_rica":[9.93,-84.08],"paris":[48.86,2.35],
}

# Canonical Scope 3 category labels (align RouteZero + summary naming)
S3_CANON = [
 (("purchased",), "1. Purchased goods & services"),
 (("capital",), "2. Capital goods"),
 (("fuel", "energy"), "3. Fuel & energy-related"),
 (("transportation and dist", "upstream transp"), "4. Upstream transport & distribution"),
 (("waste",), "5. Waste in operations"),
 (("business travel",), "6. Business travel"),
 (("commut",), "7. Employee commuting"),
 (("leased",), "8. Upstream leased assets"),
 (("use of sold",), "11. Use of sold products"),
 (("end-of-life", "end of life", "eol"), "12. End-of-life of sold products"),
]
def canon_s3(name):
    n = str(name).lower()
    for keys, label in S3_CANON:
        if any(k in n for k in keys):
            return label
    return None

# ---------------------------------------------------------------------------
data = {
    "meta": {
        "company": "Vantage Specialty Chemicals",
        "generated": str(date.today()),
        "baseline_year": 2021,
        "units": {"emissions": "tCO2e", "water": "ML", "production": "metric tons"},
        "target_type": "Internal (not SBTi-validated)",
        "scope2_default_basis": "market",
        "bau_assumption": "Held flat at last actual for 2026-2030",
    },
    "sites": [], "emissions_by_site": [], "emissions_company": {},
    "scope3_categories": {}, "water_company": {}, "water_by_site": [],
    "production_by_site": [], "production_company": {},
    "targets": {"ghg_s1s2": [], "water": []}, "initiatives": [],
}

for row in SITE_MASTER:
    sid, name, country, region, ftype, status, stress, risk, producing, _ = row
    data["sites"].append({
        "site_id": sid, "name": name, "country": country, "region": region,
        "facility_type": ftype, "status": status,
        "wwf_water_stress": stress or "NA", "wwf_water_risk": risk or "NA",
        "producing": producing,
        "lat": COORDS.get(sid, [None, None])[0], "lon": COORDS.get(sid, [None, None])[1],
    })

# ---- Emissions by site 2021-2023  (CDP :: Data 2021 to 2023) ----------------
wb = load("cdp"); ws = wb["Data 2021 to 2023"]
for r in rows(ws):
    sid = site_id(r[0]) if r and r[0] else None
    if not sid: continue
    # cols: 0 name | 1 S1_21 | 2 S2_21 | 3 sum | 4 S1_22 | 5 S2_22 | 6 sum | 7 prod22 | 8 S1_23 | 9 S2_23
    for yr, c1, c2 in [(2021,1,2),(2022,4,5),(2023,8,9)]:
        s1, s2 = num(r[c1]) if len(r)>c1 else None, num(r[c2]) if len(r)>c2 else None
        if s1 is None and s2 is None: continue
        data["emissions_by_site"].append({"site_id": sid, "year": yr,
            "scope1": round(s1 or 0, 2), "scope2_market": round(s2 or 0, 2)})
wb.close()

# ---- Emissions by site 2024 & 2025  (Results by Site) -----------------------
def extract_results_by_site(key, year):
    """Results-by-Site sheets have a leading blank column; locate 'Site Name'."""
    wb = load(key); ws = wb["Results by Site"]
    rr = rows(ws, maxc=6)
    hdr = ncol = None
    for i, r in enumerate(rr):
        for j, c in enumerate(r):
            if isinstance(c, str) and c.strip() == "Site Name":
                hdr, ncol = i, j; break
        if hdr is not None: break
    if hdr is None: wb.close(); return
    for r in rr[hdr+1:]:
        nm = r[ncol] if len(r) > ncol else None
        if nm and norm(nm) in ("grand total","total"): break   # end of the emissions table
        if not nm: continue
        sid = site_id(nm)
        if not sid: continue
        s1 = num(r[ncol+1]) if len(r) > ncol+1 else None
        s2 = num(r[ncol+2]) if len(r) > ncol+2 else None
        if s1 is None and s2 is None: continue
        if (s1 and abs(s1) > 1e7) or (s2 and abs(s2) > 1e7): continue  # guard vs usage cols
        data["emissions_by_site"].append({"site_id": sid, "year": year,
            "scope1": round(s1 or 0,2), "scope2_market": round(s2 or 0,2)})
    wb.close()

extract_results_by_site("fy24", 2024)
extract_results_by_site("fy25", 2025)

# ---- Company emissions per year (rollup S1/S2 market + S2 location + S3) -----
comp = {}
for e in data["emissions_by_site"]:
    y = e["year"]; comp.setdefault(y, {"scope1":0.0,"scope2_market":0.0})
    comp[y]["scope1"] += e["scope1"]; comp[y]["scope2_market"] += e["scope2_market"]
for y in comp:
    comp[y]["scope1"] = round(comp[y]["scope1"],1)
    comp[y]["scope2_market"] = round(comp[y]["scope2_market"],1)
# Scope 2 location-based (company) per year, from inventories/RouteZero
S2_LOCATION = {2021: 46272.5, 2022: 41478.2, 2023: 35782.6, 2024: 39042.7, 2025: 31767.6}
for y, v in S2_LOCATION.items():
    comp.setdefault(y, {}).update({"scope2_location": v})
comp.setdefault(2023,{}).update({"scope3_total": round(976825.45-57632.51-31315.12,1)})
comp.setdefault(2024,{}).update({"scope3_total": round(891694.54-59695.38-22562.57,1)})

# Per-site Scope 2 location-based: scale market by the company location/market ratio
# (site-level location factors aren't reported; this preserves the exact company total).
for e in data["emissions_by_site"]:
    cm = comp.get(e["year"], {})
    ratio = (cm["scope2_location"]/cm["scope2_market"]) if cm.get("scope2_market") else 1.0
    e["scope2_location"] = round(e["scope2_market"]*ratio, 2)

data["emissions_company"] = {str(y): comp[y] for y in sorted(comp)}

# ---- Scope 3 categories 2023 & 2024  (FY24 summary Total Inventory Summary) --
wb = load("fy24s"); ws = wb["Total Inventory Summary"]
cat24, cat23 = {}, {}
for r in rows(ws, maxc=9):
    name = r[2] if len(r)>2 else None          # category name is in column C
    val24 = num(r[3]) if len(r)>3 else None     # FY24 emissions
    val23 = num(r[7]) if len(r)>7 else None     # FY23 emissions
    lab = canon_s3(name) if isinstance(name, str) else None
    if lab and val24 is not None:
        cat24[lab] = round(val24,1)
        if val23 is not None: cat23[lab] = round(val23,1)
data["scope3_categories"] = {"2023": cat23, "2024": cat24}

# ---- Scope 3 categories 2021 & 2022  (RouteZero 'emissions' sheet) -----------
wb_rz = openpyxl.load_workbook(os.path.join(ROOT, FILES["rz"]), read_only=True, data_only=True)
ws_rz = wb_rz["emissions"]
s3 = {2021: {}, 2022: {}}
for r in ws_rz.iter_rows(min_row=2, values_only=True):
    if r[22] != "Scope 3":            # col 22 = Scope
        continue
    yr = int(r[16]) if r[16] else None  # col 16 = Reporting Period Year
    lab = canon_s3(r[23])               # col 23 = Scope Category
    mkt = r[41] or 0                    # col 41 = Market tCO2e
    if yr in s3 and lab:
        s3[yr][lab] = round(s3[yr].get(lab, 0) + mkt, 1)
wb_rz.close()
for y in (2021, 2022):
    data["scope3_categories"][str(y)] = s3[y]
    comp.setdefault(y, {}).update({"scope3_total": round(sum(s3[y].values()), 1)})
data["emissions_company"] = {str(y): comp[y] for y in sorted(comp)}
wb.close()

# ---- Water: company per year (CDP Target Tracking) + 2025 from FY25 gallons --
wb = load("cdp"); ws = wb["Target Tracking"]
water_rows, mode = [], None
for r in rows(ws, maxc=5):
    if r and isinstance(r[0],str) and r[0].strip()=="Year" and r[1] and "Water" in str(r[1]):
        mode = "water"; continue
    if mode=="water":
        y = num(r[0])
        if y is None: mode=None; continue
        if 2021 <= y <= 2025:
            data["water_company"][str(int(y))] = {"withdrawal_ml": round(num(r[1]),1)}
wb.close()
# 2025 actual from FY25 inventory (gallons -> ML), overrides planning figure
wb = load("fy25"); ws = wb["Pivot"]
for r in rows(ws, maxc=12):
    for c in r:
        if isinstance(c,(int,float)) and c and c > 1e8:   # the big gallons total
            data["water_company"]["2025"] = {"withdrawal_ml": round(c*GAL_TO_ML,1),
                                              "note": "Converted from FY25 inventory gallons"}
            break
wb.close()

# ---- Water by site 2024 (CDP Summary Water 2024) ----------------------------
wb = load("cdp"); ws = wb["Summary Water 2024"]
for r in rows(ws, maxc=14):
    nm = r[0] if r else None
    if not nm or norm(nm) in ("site","totals","total"): continue
    sid = site_id(nm)
    if not sid: continue
    surface = num(r[5]) or 0; ground = num(r[6]) or 0; third = num(r[7]) or 0; recyc = num(r[8]) or 0
    withdrawal = surface+ground+third+recyc
    consumption = num(r[12])
    if withdrawal == 0 and not consumption: continue
    basin = r[2] if len(r)>2 and isinstance(r[2],str) else None
    data["water_by_site"].append({"site_id": sid, "year": 2024,
        "withdrawal_ml": round(withdrawal,2),
        "consumption_ml": round(consumption,2) if consumption is not None else None,
        "river_basin": (basin if basin and basin.upper()!="NA" else None),
        "surface": round(surface,2),"groundwater": round(ground,2),
        "third_party": round(third,2),"recycled": round(recyc,2)})
wb.close()

# ---- Production by site 2022-2025 (Production Volumes :: Prod Volumes) -------
wb = load("prod"); ws = wb["Prod Volumes"]
rr = rows(ws, maxc=15)
cur_year = None
for r in rows(ws, maxc=15):
    # year header rows look like [None, 2022, ...]
    for c in r[:3]:
        if isinstance(c,(int,float)) and 2021 <= c <= 2030:
            cur_year = int(c)
    nm = r[0] if r else None
    if not nm or not isinstance(nm,str): continue
    if norm(nm) in ("production volume","total"): continue
    sid = site_id(nm)
    if not sid or cur_year is None or cur_year > 2025: continue   # 2026 = partial budget, exclude
    tons = num(r[14]) if len(r)>14 else None   # "Tons <year>" last col
    if tons is None: continue
    data["production_by_site"].append({"site_id": sid, "year": cur_year, "production_tons": round(tons,1)})
wb.close()
pc = {}
for p in data["production_by_site"]:
    pc[p["year"]] = pc.get(p["year"],0)+p["production_tons"]
data["production_company"] = {str(y): round(pc[y],0) for y in sorted(pc)}

# ---- Targets (CDP Target Tracking) ------------------------------------------
wb = load("cdp"); ws = wb["Target Tracking"]
mode = None
for r in rows(ws, maxc=5):
    h = r[0] if r else None
    if isinstance(h,str) and h.strip()=="Year" and r[1] and "GHG" in str(r[1]): mode="ghg"; continue
    if isinstance(h,str) and h.strip()=="Year" and r[1] and "Water" in str(r[1]): mode="water"; continue
    y = num(h)
    if y is None or not (2021 <= y <= 2030):
        if isinstance(h,str) and h.strip() and h.strip()!="Year": mode=None
        continue
    if mode=="ghg":
        data["targets"]["ghg_s1s2"].append({"year":int(y),"actual":num(r[1]),"target":num(r[2])})
    elif mode=="water":
        data["targets"]["water"].append({"year":int(y),"actual":num(r[1]),
            "target_25":num(r[2]),"target_50":num(r[3])})
wb.close()

# Override target "actuals" with inventory-derived actuals (the planning sheet's
# 2025 was a pre-close forecast). Keeps target lines; future actuals -> null.
ghg_actual = {y: round(comp[y]["scope1"]+comp[y]["scope2_market"])
              for y in comp if "scope1" in comp[y] and "scope2_market" in comp[y]}
for t in data["targets"]["ghg_s1s2"]:
    t["actual"] = ghg_actual.get(t["year"]) if t["year"] <= 2025 else None
for t in data["targets"]["water"]:
    w = data["water_company"].get(str(t["year"]))
    t["actual"] = (w["withdrawal_ml"] if w else None) if t["year"] <= 2025 else None

# ---- Initiatives (CDP GHG-Water Initiatives (2)) ----------------------------
wb = load("cdp"); ws = wb["GHG-Water Initiatives (2)"]
rr = rows(ws, maxc=14)
hdr = None
for i,r in enumerate(rr):
    if r and isinstance(r[0],str) and r[0].strip()=="Project": hdr=i; break
if hdr is not None:
    for r in rr[hdr+1:]:
        if not r or not r[0]: continue
        proj = str(r[0]).strip()
        if not proj or norm(proj) in ("total","grand total"): continue
        def g(i): return r[i] if len(r)>i else None
        data["initiatives"].append({
            "name": proj, "site_id": site_id(g(1)) or norm(g(1)),
            "site_name": str(g(1)).strip() if g(1) else "",
            "stage": str(g(2)).strip() if g(2) else "",
            "year": int(num(g(3))) if num(g(3)) else None,
            "type": str(g(4)).strip() if g(4) else "",
            "ghg_savings_tco2e": num(g(5)), "scope1": num(g(6)), "scope2": num(g(7)),
            "water_savings_ml": num(g(8)), "energy_savings_mwh": num(g(9)),
            "incentive_usd": num(g(10)), "capex_usd": num(g(11)),
            "opex_savings_usd": num(g(12)), "payback_years": num(g(13)),
        })
wb.close()

# ---------------------------------------------------------------------------
out = os.path.join(ROOT, "data.json")
with open(out, "w") as f:
    json.dump(data, f, indent=1, ensure_ascii=False)

# ---- Inject data + logo into the template -> self-contained dashboard.html --
import base64
tpl_path = os.path.join(ROOT, "build", "dashboard_template.html")
if os.path.exists(tpl_path):
    with open(tpl_path, encoding="utf-8") as f:
        html = f.read()
    logo_path = os.path.join(ROOT, "assets", "vantage-logo.jpg")
    if os.path.exists(logo_path):
        b64 = base64.b64encode(open(logo_path, "rb").read()).decode()
        html = html.replace("__LOGO_DATA_URI__", "data:image/jpeg;base64," + b64)
    wm_path = os.path.join(ROOT, "build", "worldmap.txt")
    if os.path.exists(wm_path):
        html = html.replace("__WORLDMAP_PATH__", open(wm_path, encoding="utf-8").read())
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    html = html.replace("const D = window.DASHBOARD_DATA;",
                        "window.DASHBOARD_DATA=" + payload + ";\nconst D = window.DASHBOARD_DATA;")
    dash = os.path.join(ROOT, "dashboard.html")
    with open(dash, "w", encoding="utf-8") as f:
        f.write(html)
    print("=== dashboard.html written:", f"{os.path.getsize(dash)//1024} KB ===")

# ---- Validation report ------------------------------------------------------
def co(y, k): return data["emissions_company"].get(str(y),{}).get(k)
print("=== data.json written:", out, f"({os.path.getsize(out)//1024} KB) ===\n")
print("Sites:", len(data['sites']), "| emissions rows:", len(data['emissions_by_site']),
      "| water-site rows:", len(data['water_by_site']),
      "| production rows:", len(data['production_by_site']), "| initiatives:", len(data['initiatives']))
print("\nScope 1+2 (market) company totals vs anchors:")
anchors = {2021:107249,2022:106088,2023:88585,2024:82258,2025:80144}
for y in range(2021,2026):
    s1,s2 = co(y,"scope1"), co(y,"scope2_market")
    tot = round((s1 or 0)+(s2 or 0))
    a = anchors[y]; flag = "OK" if abs(tot-a) <= max(50, a*0.01) else "CHECK"
    print(f"  {y}: S1={s1}  S2={s2}  S1+2={tot}  (anchor {a})  [{flag}]")
print("\nWater company (ML):", {k:v.get('withdrawal_ml') for k,v in data['water_company'].items()})
print("Production company (tons):", data['production_company'])
print("Scope 3 2024 categories:", len(data['scope3_categories']['2024']),
      "sum=", round(sum(data['scope3_categories']['2024'].values())))
print("Initiatives stages:", sorted({i['stage'] for i in data['initiatives']}))
if UNMAPPED:
    print("\n!! UNMAPPED site names (add to SITE_MASTER aliases):")
    for u in sorted(UNMAPPED): print("   -", repr(u))
else:
    print("\nAll site names mapped cleanly.")
